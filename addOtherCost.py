import time
from pathlib import Path

import openpyxl
from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.support.select import Select

currentPath = str(Path.cwd())
currentPath = currentPath.replace('\\', '/')
filePath = currentPath + '/' + "addOtherCost.xlsx"
wb = openpyxl.load_workbook(filePath)
sheet_to_focus = 'addOtherCost'
for s in range(len(wb.sheetnames)):
    if wb.sheetnames[s] == sheet_to_focus:
        break
wb.active = s

driver = webdriver.Chrome()
driver.get("https://test.rasatpa.ir/sso/login")
driver.maximize_window()
app = driver.find_element(By.ID, 'app-select')
username = driver.find_element(By.ID, 'username')
password = driver.find_element(By.ID, 'password')
login = driver.find_element(By.ID, 'lbtn')
time.sleep(1)
appName = "کارتابل"
timeOut = 20
app.send_keys(appName)
time.sleep(1)
username.send_keys("1270138758")
password.send_keys("Ras@4321")
time.sleep(1)
login.click()
time.sleep(3)
action = webdriver.ActionChains(driver)
try:
    organizationSelect = driver.find_element(By.XPATH,
                                             "/html/body/div[1]/div[2]/div/div/div/div/div/div/div[2]/div/a/p")
    organizationSelect.click()
except NoSuchElementException:
    pass
time.sleep(5)
try:
    collapseBtnTaskSearchPanelId = driver.find_element(By.ID, "collapse-btn-taskSearchPanelId")
    collapseBtnTaskSearchPanelId.click()
except NoSuchElementException:
    pass
time.sleep(10)
try:
    barcode = driver.find_element(By.XPATH,
                                  "/html/body/main/div/div/div/div[3]/div/div[2]/div/div[1]/div/div[2]/div["
                                  "2]/div/div/div/form/div/div[2]/div/input")
    barcode.send_keys("1402/38/34669/031/11573")
except NoSuchElementException:
    pass
time.sleep(5)
try:
    taskListGridSearchBtn = driver.find_element(By.ID, 'taskListGrid-search-btn')
    taskListGridSearchBtn.click()
except NoSuchElementException:
    pass
time.sleep(5)
try:
    record = driver.find_element(By.XPATH,
                                 "/html/body/main/div/div/div/div[3]/div/div[2]/div/div[1]/div/div[3]/div["
                                 "2]/div/div/div/div/div[1]/table/tbody/tr/td[4]/a")
    record.click()
except NoSuchElementException:
    pass
time.sleep(5)
driver.switch_to.window(driver.window_handles[1])
driver.maximize_window()
time.sleep(5)
try:
    searchClaim = driver.find_element(By.XPATH,
                                      "/html/body/main/div/div/div/div[1]/div[2]/div[2]/div/div[2]/div/div["
                                      "1]/div/div[11]/div[1]/div[2]/button")
    searchClaim.click()
except NoSuchElementException:
    pass
time.sleep(5)
try:
    claimNumber = driver.find_element(By.XPATH,
                                      "/html/body/main/div/div/div/div[1]/div[2]/div[2]/div/div[2]/div/div[1]/div/div["
                                      "11]/div[2]/div/div/div/form/div/div[1]/div/input")
    claimNumber.send_keys("1296834905")
except NoSuchElementException:
    pass
time.sleep(5)
try:
    claimGridSearchBtn = driver.find_element(By.ID, 'claimGrid-search-btn')
    claimGridSearchBtn.click()
except NoSuchElementException:
    pass
time.sleep(5)
try:
    moneyBtn = driver.find_element(By.XPATH,
                                   "/html/body/main/div/div/div/div[1]/div[2]/div[2]/div/div[2]/div/div[1]/div/div["
                                   "14]/div[2]/div/div/div/div/div[2]/table/tbody/tr/td[20]/a")
    driver.implicitly_wait(10)
    ActionChains(driver).move_to_element(moneyBtn).click(moneyBtn).perform()
except NoSuchElementException:
    pass
time.sleep(5)
try:
    driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
    addOther = driver.find_element(By.XPATH,
                                   "/html/body/main/div/div/div/div[1]/div[2]/div[2]/div/div[2]/div/div[1]/div/div["
                                   "17]/div/div/div[2]/div[2]/div[3]/div[2]/div/div/div/div/div[1]/a[2]")
    addOther.click()

except NoSuchElementException:
    pass
time.sleep(5)
try:
    selectItemContainer = Select(driver.find_element(By.ID,
                                                     "itemId"))
    selectItemContainer.select_by_visible_text("سایر")
except NoSuchElementException:
    pass
time.sleep(5)
try:
    otherTotalAmountInputId = driver.find_element(By.ID,
                                                  "otherTotalAmountInputId")
    otherTotalAmountInputId.send_keys("4000")
except NoSuchElementException:
    pass
time.sleep(5)
try:
    otherDescriptionInputId = driver.find_element(By.ID,
                                                  "otherDescriptionInputId")
    otherDescriptionInputId.send_keys("پرداخت از  محل صندوق")
except NoSuchElementException:
    pass
time.sleep(5)
try:
    saveOtherBtn = driver.find_element(By.ID, 'btnSaveother')
    saveOtherBtn.click()
except NoSuchElementException:
    pass
time.sleep(5)
try:
    claimAssessmentBtn = driver.find_element(By.ID, 'btnClaimAssessment')
    claimAssessmentBtn.click()
except NoSuchElementException:
    pass
time.sleep(5)
